"""
Excel書き込み: openpyxl

openpyxl はテンプレートファイルを直接開いて値を書き込むため、
書式（結合セル・罫線・フォント・列幅・行高・配置）が完全に保持される。
xlutils.copy のような再構築は行わない。
"""

import io
import os
import logging
from datetime import datetime
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import numbers

from transcription.cell_mappings import CELL_MAPPINGS, SUPPORTED_SHEETS, DATE_FIELDS

logger = logging.getLogger(__name__)

# ローカル: backend/templates/  デプロイ: /data/（Fly.io Volume）
_LOCAL_TEMPLATE = os.path.join(
    os.path.dirname(__file__), "..", "templates", "template.xlsx"
)
_VOLUME_TEMPLATE = "/data/template.xlsx"


def _get_template_path() -> str:
    """テンプレートファイルのパスを返す。Volume優先、なければローカル。"""
    if os.path.exists(_VOLUME_TEMPLATE):
        return _VOLUME_TEMPLATE
    return _LOCAL_TEMPLATE

# 日本語日付フォーマット
_JAPANESE_DATE_FORMAT = 'yyyy"年"m"月"d"日"'


def _parse_date(date_str: str) -> datetime | None:
    """
    日付文字列を datetime に変換。
    対応フォーマット: 2000年5月5日, 2000/5/5, 令和2年5月5日, 昭和26年5月12日 等
    """
    # 西暦
    for fmt in ["%Y年%m月%d日", "%Y/%m/%d", "%Y-%m-%d"]:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue

    # 和暦→西暦
    era_map = {
        "令和": 2018, "平成": 1988, "昭和": 1925, "大正": 1911, "明治": 1867,
    }
    for era, base_year in era_map.items():
        if era in date_str:
            try:
                cleaned = date_str.replace(era, "").strip()
                parts = cleaned.replace("年", "/").replace("月", "/").replace("日", "").split("/")
                year = base_year + int(parts[0])
                month = int(parts[1])
                day = int(parts[2])
                return datetime(year, month, day)
            except (ValueError, IndexError):
                continue

    return None


def write_to_template(
    sheet_name: str,
    fields: list[dict[str, str]],
) -> bytes:
    """
    テンプレートにデータを書き込み、.xlsx バイナリを返す。

    openpyxl で直接開いて値だけ書き込むため、書式は完全に保持される。

    Args:
        sheet_name: 書き込み先のシート名
        fields: [{"field_name": "...", "value": "..."}, ...]

    Returns:
        .xlsx ファイルのバイナリデータ
    """
    if sheet_name not in SUPPORTED_SHEETS:
        raise ValueError(f"未対応のシート: {sheet_name}（対応: {SUPPORTED_SHEETS}）")

    template_path = _get_template_path()
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"テンプレートが見つかりません: {template_path}\n"
            "backend/templates/ に (者）★原本.xlsx を配置してください。"
        )

    # テンプレートを直接開く（書式はそのまま保持される）
    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    mapping = CELL_MAPPINGS.get(sheet_name, {})
    written_count = 0

    for field in fields:
        name = field["field_name"]
        value = field["value"]

        if not value:
            continue

        if name not in mapping:
            logger.warning(f"マッピングにないフィールド: {name}")
            continue

        row, col = mapping[name]
        cell = ws.cell(row=row, column=col)

        # 日付フィールド: datetime値 + 日付フォーマット
        if name in DATE_FIELDS:
            dt = _parse_date(value)
            if dt is not None:
                cell.value = dt
                cell.number_format = _JAPANESE_DATE_FORMAT
                written_count += 1
                logger.info(f"  書き込み(日付): {name} = '{value}' → ({row},{col})")
                continue

        # テキスト: 値だけ書き込み（書式は元のまま）
        cell.value = value
        written_count += 1
        logger.info(f"  書き込み: {name} = '{value}' → ({row},{col})")

    logger.info(f"シート '{sheet_name}' に {written_count} フィールドを書き込みました")

    # メモリ上に保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
